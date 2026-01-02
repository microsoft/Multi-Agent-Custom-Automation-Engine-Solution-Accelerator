"""
GitHub Excel Generator - Utility to generate Excel reports from GitHub repository data.

This module provides functionality to fetch branch and pull request information from
a GitHub repository and generate an Excel spreadsheet with the data.
"""

import logging
import os
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from github import Github, GithubException

logger = logging.getLogger(__name__)


class GitHubExcelGenerator:
    """Generator for Excel reports containing GitHub repository branch and PR information."""

    def __init__(self, github_token: Optional[str] = None):
        """
        Initialize the GitHub Excel Generator.

        Args:
            github_token: GitHub personal access token. If not provided, will attempt
                         to read from GITHUB_TOKEN environment variable.
        """
        self.token = github_token or os.getenv("GITHUB_TOKEN")
        self.github_client = None
        if self.token:
            try:
                self.github_client = Github(self.token)
                logger.info("GitHub client initialized successfully")
            except Exception as e:
                logger.error(f"Failed to initialize GitHub client: {e}")
        else:
            logger.warning("No GitHub token provided. Some features may not work.")

    def get_branch_info(self, owner: str, repo: str) -> List[Dict]:
        """
        Fetch branch information from a GitHub repository.

        Args:
            owner: Repository owner (username or organization)
            repo: Repository name

        Returns:
            List of dictionaries containing branch information
        """
        if not self.github_client:
            logger.error("GitHub client not initialized")
            return []

        branch_info = []
        try:
            repository = self.github_client.get_repo(f"{owner}/{repo}")
            branches = repository.get_branches()

            for branch in branches:
                pr_status = "none"
                pr_url = None

                # Check for associated pull requests
                try:
                    pulls = repository.get_pulls(state="all", head=f"{owner}:{branch.name}")
                    pull_list = list(pulls)
                    
                    if pull_list:
                        # Get the most recent PR for this branch
                        latest_pr = pull_list[0]
                        if latest_pr.merged:
                            pr_status = "merged"
                        elif latest_pr.state == "open":
                            pr_status = "open"
                        elif latest_pr.state == "closed":
                            pr_status = "closed"
                        pr_url = latest_pr.html_url
                except Exception as pr_error:
                    logger.debug(f"Error fetching PRs for branch {branch.name}: {pr_error}")

                # Get branch creator from the first commit
                created_by = "Unknown"
                try:
                    commit = branch.commit
                    if commit and commit.author:
                        created_by = commit.author.login
                    elif commit and commit.commit and commit.commit.author:
                        created_by = commit.commit.author.name
                except Exception as e:
                    logger.debug(f"Error getting creator for branch {branch.name}: {e}")

                branch_info.append({
                    "branch_name": branch.name,
                    "pr_status": pr_status,
                    "created_by": created_by,
                    "pr_url": pr_url or "N/A",
                    "protected": branch.protected,
                    "last_commit_date": commit.commit.author.date.strftime("%Y-%m-%d %H:%M:%S") if commit and commit.commit and commit.commit.author else "N/A"
                })

            logger.info(f"Successfully fetched information for {len(branch_info)} branches")
            return branch_info

        except GithubException as e:
            logger.error(f"GitHub API error: {e}")
            return []
        except Exception as e:
            logger.error(f"Error fetching branch information: {e}")
            return []

    def generate_excel(self, branch_data: List[Dict], output_path: str) -> bool:
        """
        Generate an Excel file from branch data.

        Args:
            branch_data: List of dictionaries containing branch information
            output_path: Path where the Excel file should be saved

        Returns:
            True if successful, False otherwise
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Branch Information"

            # Define headers
            headers = [
                "Branch Name",
                "PR Status",
                "Created By",
                "PR URL",
                "Protected",
                "Last Commit Date"
            ]

            # Style for headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Write data
            for row_num, branch in enumerate(branch_data, 2):
                ws.cell(row=row_num, column=1).value = branch.get("branch_name", "")
                ws.cell(row=row_num, column=2).value = branch.get("pr_status", "none")
                ws.cell(row=row_num, column=3).value = branch.get("created_by", "Unknown")
                ws.cell(row=row_num, column=4).value = branch.get("pr_url", "N/A")
                ws.cell(row=row_num, column=5).value = "Yes" if branch.get("protected") else "No"
                ws.cell(row=row_num, column=6).value = branch.get("last_commit_date", "N/A")

                # Apply alternating row colors for better readability
                if row_num % 2 == 0:
                    fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=col).fill = fill

            # Adjust column widths
            column_widths = [25, 15, 25, 60, 12, 20]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_num)].width = width

            # Save the workbook
            wb.save(output_path)
            logger.info(f"Excel file successfully created at {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error generating Excel file: {e}")
            return False

    def generate_report(self, owner: str, repo: str, output_path: str) -> bool:
        """
        Generate a complete Excel report for a GitHub repository.

        Args:
            owner: Repository owner
            repo: Repository name
            output_path: Path where the Excel file should be saved

        Returns:
            True if successful, False otherwise
        """
        logger.info(f"Generating report for {owner}/{repo}")
        branch_data = self.get_branch_info(owner, repo)
        
        if not branch_data:
            logger.warning("No branch data found or error occurred")
            return False
        
        return self.generate_excel(branch_data, output_path)
