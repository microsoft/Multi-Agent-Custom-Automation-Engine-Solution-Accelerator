"""
GitHub Excel Generator using MCP Tools - Utility to generate Excel reports from GitHub repository data.

This module provides functionality to fetch branch and pull request information from
a GitHub repository using the GitHub MCP server tools and generate an Excel spreadsheet.
"""

import logging
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)


class GitHubExcelGeneratorMCP:
    """Generator for Excel reports containing GitHub repository branch and PR information using MCP tools."""

    def __init__(self, github_tools):
        """
        Initialize the GitHub Excel Generator with MCP tools.

        Args:
            github_tools: GitHub MCP tools interface for making API calls
        """
        self.github_tools = github_tools
        logger.info("GitHub Excel Generator MCP initialized")

    def format_branch_data(self, branches: List[Dict], prs: List[Dict]) -> List[Dict]:
        """
        Format branch and PR data for Excel export.

        Args:
            branches: List of branch dictionaries from GitHub API
            prs: List of pull request dictionaries from GitHub API

        Returns:
            List of formatted dictionaries containing branch information
        """
        # Create a mapping of branch names to PRs
        pr_by_branch = {}
        for pr in prs:
            head_ref = pr.get('head', {}).get('ref')
            if head_ref:
                # Keep the most recent PR for each branch
                if head_ref not in pr_by_branch:
                    pr_by_branch[head_ref] = pr

        branch_info = []
        for branch in branches:
            branch_name = branch.get('name', 'Unknown')
            pr = pr_by_branch.get(branch_name)
            
            pr_status = "none"
            pr_url = "N/A"
            created_by = "Unknown"
            
            if pr:
                # Determine PR status
                if pr.get('merged', False):
                    pr_status = "merged"
                elif pr.get('state') == 'open':
                    pr_status = "open"
                elif pr.get('state') == 'closed':
                    pr_status = "closed"
                
                pr_url = pr.get('html_url', 'N/A')
                created_by = pr.get('user', {}).get('login', 'Unknown')
            
            # If no PR, try to get branch creator (this would require commit info)
            # For now, we'll use the PR creator if available
            
            branch_info.append({
                "branch_name": branch_name,
                "pr_status": pr_status,
                "created_by": created_by,
                "pr_url": pr_url,
                "protected": branch.get('protected', False),
            })

        logger.info(f"Formatted information for {len(branch_info)} branches")
        return branch_info

    def generate_excel(self, branch_data: List[Dict], output_path: str) -> bool:
        """
        Generate an Excel file from branch data.

        Args:
            branch_data: List of dictionaries containing branch information
            output_path: Path where the Excel file should be saved

        Returns:
            True if successful, False otherwise
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Branch Information"

            # Define headers
            headers = [
                "Branch Name",
                "PR Status",
                "Created By",
                "PR URL",
                "Protected",
            ]

            # Style for headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Write data
            for row_num, branch in enumerate(branch_data, 2):
                ws.cell(row=row_num, column=1).value = branch.get("branch_name", "")
                ws.cell(row=row_num, column=2).value = branch.get("pr_status", "none")
                ws.cell(row=row_num, column=3).value = branch.get("created_by", "Unknown")
                ws.cell(row=row_num, column=4).value = branch.get("pr_url", "N/A")
                ws.cell(row=row_num, column=5).value = "Yes" if branch.get("protected") else "No"

                # Apply alternating row colors for better readability
                if row_num % 2 == 0:
                    fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=col).fill = fill

            # Adjust column widths
            column_widths = [25, 15, 25, 60, 12]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_num)].width = width

            # Save the workbook
            wb.save(output_path)
            logger.info(f"Excel file successfully created at {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error generating Excel file: {e}")
            return False
