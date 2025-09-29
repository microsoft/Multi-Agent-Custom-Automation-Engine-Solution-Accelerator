"""Shared utilities for working with uploaded tabular datasets."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import load_workbook


ALLOWED_EXTENSIONS = {".csv", ".xlsx"}


def ensure_allowed_extension(path: Path) -> str:
    """Validate that the dataset extension is supported and return it."""

    extension = path.suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        allowed = ", ".join(sorted(ALLOWED_EXTENSIONS))
        raise ValueError(
            f"Unsupported file extension '{extension}'. Supported extensions: {allowed}"
        )
    return extension


def _sanitize_columns(columns: Iterable[Any]) -> List[str]:
    """Normalize column names to strings without empty values."""

    normalized: List[str] = []
    for idx, column in enumerate(columns):
        if column is None:
            normalized.append(f"column_{idx + 1}")
        else:
            value = str(column).strip()
            normalized.append(value or f"column_{idx + 1}")
    return normalized


def read_preview(path: Path, max_rows: int = 5) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Return column names and the first *max_rows* rows for display purposes."""

    extension = ensure_allowed_extension(path)

    if extension == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            columns = _sanitize_columns(reader.fieldnames or [])
            rows: List[Dict[str, Any]] = []

            for raw in reader:
                row = {col: raw.get(col, "") for col in columns}
                rows.append(row)
                if len(rows) >= max_rows:
                    break

    else:  # .xlsx
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)

        header = next(rows_iter, [])
        columns = _sanitize_columns(header)
        rows = []

        for raw in rows_iter:
            row = {columns[idx]: raw[idx] for idx in range(min(len(columns), len(raw)))}
            for key, value in row.items():
                if isinstance(value, datetime):
                    row[key] = value.isoformat()
                elif value is None:
                    row[key] = ""
            rows.append(row)
            if len(rows) >= max_rows:
                break

        workbook.close()

    return columns, rows


def iter_tabular_records(path: Path) -> Iterable[Dict[str, Any]]:
    """Yield row dictionaries for the given dataset."""

    extension = ensure_allowed_extension(path)

    if extension == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            columns = _sanitize_columns(reader.fieldnames or [])
            for raw in reader:
                yield {col: raw.get(col, "") for col in columns}

    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        header = next(rows_iter, [])
        columns = _sanitize_columns(header)

        for raw in rows_iter:
            row = {columns[idx]: raw[idx] for idx in range(min(len(columns), len(raw)))}
            yield row

        workbook.close()


def extract_numeric_series(
    path: Path, target_column: str
) -> Tuple[List[float], int]:
    """Extract numeric values from a column, returning the series and total row count."""

    values: List[float] = []
    row_count = 0

    for record in iter_tabular_records(path):
        row_count += 1
        raw_value = record.get(target_column)
        if raw_value is None or raw_value == "":
            continue
        try:
            values.append(float(raw_value))
        except (TypeError, ValueError):
            continue

    if not values:
        raise ValueError(
            f"Column '{target_column}' does not contain numeric data suitable for forecasting"
        )

    return values, row_count


def detect_numeric_columns(path: Path, sample_limit: int = 50) -> List[str]:
    """Infer which columns appear to contain numeric values."""

    numeric_candidates: Dict[str, int] = {}
    examined = 0

    for record in iter_tabular_records(path):
        examined += 1
        for key, value in record.items():
            if key not in numeric_candidates:
                numeric_candidates[key] = 0
            if value in (None, ""):
                continue
            try:
                float(value)
            except (TypeError, ValueError):
                continue
            else:
                numeric_candidates[key] += 1

        if examined >= sample_limit:
            break

    return [
        column
        for column, numeric_count in numeric_candidates.items()
        if numeric_count > 0
    ]


def simple_linear_forecast(values: List[float], periods: int = 3) -> List[float]:
    """Project future values using a simple linear regression over the index."""

    if len(values) < 2:
        raise ValueError("Need at least two data points to generate a forecast")

    n = len(values)
    x_values = list(range(n))

    sum_x = sum(x_values)
    sum_y = sum(values)
    sum_xy = sum(x * y for x, y in zip(x_values, values))
    sum_x2 = sum(x * x for x in x_values)

    denominator = (n * sum_x2) - (sum_x ** 2)
    if denominator == 0:
        # Fallback to repeating the last value
        return [values[-1] for _ in range(periods)]

    slope = ((n * sum_xy) - (sum_x * sum_y)) / denominator
    intercept = (sum_y - slope * sum_x) / n

    forecast = []
    for step in range(1, periods + 1):
        future_index = n - 1 + step
        forecast.append(slope * future_index + intercept)

    return forecast


def summarize_numeric_series(values: List[float]) -> Dict[str, Any]:
    """Return simple descriptive statistics for a numeric series."""

    if not values:
        return {"count": 0}

    return {
        "count": len(values),
        "min": min(values),
        "max": max(values),
        "mean": mean(values),
    }
