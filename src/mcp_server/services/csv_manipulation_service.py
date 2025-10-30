"""CSV Data Manipulation MCP tools for creating, editing, and transforming CSV files."""

from __future__ import annotations

import csv
import json
import logging
import re
import shutil
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from core.factory import Domain, MCPToolBase

LOGGER = logging.getLogger(__name__)

METADATA_FILENAME = "metadata.json"


def _sanitize_identifier(value: str) -> str:
    """Sanitize identifier for use in file paths."""
    return re.sub(r"[^A-Za-z0-9_-]", "_", value)


class CSVManipulationService(MCPToolBase):
    """Tools for creating, editing, and manipulating CSV files."""

    def __init__(self, dataset_root: Optional[Path] = None):
        super().__init__(Domain.CSV_MANIPULATION)
        default_root = Path(__file__).resolve().parents[3] / "data" / "uploads"
        self.dataset_root = dataset_root or default_root
        self.dataset_root.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # Helpers
    def _metadata_path(self, user_id: str, dataset_id: str) -> Path:
        """Get path to metadata file."""
        user_folder = _sanitize_identifier(user_id)
        dataset_folder = _sanitize_identifier(dataset_id)
        return self.dataset_root / user_folder / dataset_folder / METADATA_FILENAME

    def _dataset_path(self, user_id: str, dataset_id: str, filename: Optional[str] = None) -> Path:
        """Get path to dataset file. Searches across all users if not found for specified user."""
        user_folder = _sanitize_identifier(user_id)
        dataset_folder = _sanitize_identifier(dataset_id)
        
        if filename:
            file_path = self.dataset_root / user_folder / dataset_folder / filename
            if file_path.exists():
                return file_path
            # Search across all users
            for user_dir in self.dataset_root.iterdir():
                if not user_dir.is_dir():
                    continue
                candidate_path = user_dir / dataset_folder / filename
                if candidate_path.exists():
                    LOGGER.info(f"Found dataset {dataset_id} under user {user_dir.name}")
                    return candidate_path
            return file_path
        
        # Try to find existing CSV file
        dataset_dir = self.dataset_root / user_folder / dataset_folder
        if dataset_dir.exists():
            for file in dataset_dir.iterdir():
                if file.suffix.lower() == '.csv':
                    return file
        
        # Search across all users
        LOGGER.info(f"Dataset {dataset_id} not found for user {user_id}, searching all users...")
        for user_dir in self.dataset_root.iterdir():
            if not user_dir.is_dir():
                continue
            candidate_dir = user_dir / dataset_folder
            if candidate_dir.exists():
                for file in candidate_dir.iterdir():
                    if file.suffix.lower() == '.csv':
                        LOGGER.info(f"Found dataset {dataset_id} under user {user_dir.name}")
                        return file
        
        return dataset_dir / "data.csv"

    def _get_metadata(self, dataset_id: str, user_id: str = "default") -> Dict[str, Any]:
        """Get metadata for a dataset. Searches across all users if not found for specified user."""
        metadata_path = self._metadata_path(user_id, dataset_id)
        
        # Try specified user first
        if metadata_path.exists():
            with open(metadata_path, "r", encoding="utf-8") as f:
                return json.load(f)
        
        # Search across all users
        LOGGER.info(f"Metadata for dataset {dataset_id} not found for user {user_id}, searching all users...")
        for user_dir in self.dataset_root.iterdir():
            if not user_dir.is_dir():
                continue
            candidate_metadata = self._metadata_path(user_dir.name, dataset_id)
            if candidate_metadata.exists():
                LOGGER.info(f"Found metadata for dataset {dataset_id} under user {user_dir.name}")
                with open(candidate_metadata, "r", encoding="utf-8") as f:
                    return json.load(f)
        
        raise FileNotFoundError(f"No metadata found for dataset '{dataset_id}'")

    def _save_metadata(self, metadata: Dict[str, Any], user_id: str) -> None:
        """Save metadata file."""
        dataset_id = metadata.get("dataset_id")
        if not dataset_id:
            raise ValueError("Metadata must include dataset_id")
        
        metadata_path = self._metadata_path(user_id, dataset_id)
        metadata_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(metadata_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f, indent=2)

    def _read_csv_data(self, file_path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Read CSV file and return headers and rows."""
        with open(file_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            rows = list(reader)
        return list(headers), rows

    def _write_csv_data(self, file_path: Path, headers: List[str], rows: List[Dict[str, Any]]) -> None:
        """Write CSV file with headers and rows."""
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(file_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)

    def _create_backup(self, file_path: Path) -> Path:
        """Create backup copy of file."""
        backup_path = file_path.with_suffix(f".backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        shutil.copy2(file_path, backup_path)
        return backup_path

    # ------------------------------------------------------------------
    # MCP Tool Registration
    @property
    def tool_count(self) -> int:
        return 10

    def register_tools(self, mcp: Any) -> None:
        """Register all CSV manipulation tools with the MCP server."""

        @mcp.tool(tags={self.domain.value})
        def create_csv_file(
            filename: str,
            columns: List[str],
            data_rows: Optional[List[List[Any]]] = None,
            user_id: str = "default"
        ) -> Dict[str, Any]:
            """
            Create a new CSV file with specified structure.
            
            Args:
                filename: Name for the CSV file (without extension)
                columns: List of column names
                data_rows: Optional list of rows (each row is a list of values matching columns)
                user_id: User identifier (default: "default")
            
            Returns:
                Dictionary with dataset_id, file path, and creation status.
            """
            try:
                # Generate unique dataset ID
                dataset_id = str(uuid.uuid4())
                
                # Ensure filename has .csv extension
                if not filename.lower().endswith('.csv'):
                    filename = f"{filename}.csv"
                
                file_path = self._dataset_path(user_id, dataset_id, filename)
                
                # Prepare rows
                rows = []
                if data_rows:
                    for row_data in data_rows:
                        if len(row_data) != len(columns):
                            raise ValueError(
                                f"Row length {len(row_data)} doesn't match columns length {len(columns)}"
                            )
                        row_dict = {col: str(val) for col, val in zip(columns, row_data)}
                        rows.append(row_dict)
                
                # Write CSV file
                self._write_csv_data(file_path, columns, rows)
                
                # Create metadata
                metadata = {
                    "dataset_id": dataset_id,
                    "original_filename": filename,
                    "stored_filename": filename,
                    "uploaded_at": datetime.now().isoformat(),
                    "size_bytes": file_path.stat().st_size,
                    "columns": columns,
                    "row_count": len(rows),
                    "created_by": "csv_manipulation_service"
                }
                
                self._save_metadata(metadata, user_id)
                
                LOGGER.info(f"Created CSV file: {file_path} with {len(rows)} rows")
                
                return {
                    "success": True,
                    "dataset_id": dataset_id,
                    "filename": filename,
                    "file_path": str(file_path),
                    "columns": columns,
                    "row_count": len(rows),
                    "message": f"CSV file '{filename}' created successfully with {len(rows)} rows"
                }
                
            except Exception as e:
                LOGGER.error(f"Error creating CSV file: {e}", exc_info=True)
                return {
                    "success": False,
                    "error": str(e)
                }

        @mcp.tool(tags={self.domain.value})
        def read_csv_file(
            dataset_id: str,
            user_id: str = "default",
            rows_limit: Optional[int] = None
        ) -> Dict[str, Any]:
            """
            Read CSV file with optional row limit.
            
            Args:
                dataset_id: The dataset ID to read
                user_id: User identifier (default: "default")
                rows_limit: Optional limit on number of rows to return
            
            Returns:
                Dictionary with columns, rows, and metadata.
            """
            try:
                file_path = self._dataset_path(user_id, dataset_id)
                
                if not file_path.exists():
                    return {
                        "success": False,
                        "error": f"CSV file not found for dataset '{dataset_id}'"
                    }
                
                headers, rows = self._read_csv_data(file_path)
                
                # Apply limit if specified
                if rows_limit and rows_limit > 0:
                    rows = rows[:rows_limit]
                
                return {
                    "success": True,
                    "dataset_id": dataset_id,
                    "columns": headers,
                    "rows": rows,
                    "row_count": len(rows),
                    "total_rows": len(rows) if rows_limit is None else None
                }
                
            except Exception as e:
                LOGGER.error(f"Error reading CSV file: {e}", exc_info=True)
                return {
                    "success": False,
                    "error": str(e)
                }

        @mcp.tool(tags={self.domain.value})
        def edit_csv_file(
            dataset_id: str,
            operations: List[Dict[str, Any]],
            user_id: str = "default"
        ) -> Dict[str, Any]:
            """
            Apply edits to CSV file (add/update/delete rows, transform columns).
            
            Args:
                dataset_id: The dataset ID to edit
                user_id: User identifier (default: "default")
                operations: List of operations to perform. Each operation is a dict with:
                    - type: "add_row", "update_row", "delete_row", "rename_column", 
                             "add_column", "remove_column", "sort"
                    - For add_row: {"type": "add_row", "row": {...}}
                    - For update_row: {"type": "update_row", "index": int, "row": {...}}
                    - For delete_row: {"type": "delete_row", "index": int}
                    - For rename_column: {"type": "rename_column", "old_name": str, "new_name": str}
                    - For add_column: {"type": "add_column", "name": str, "default_value": Any}
                    - For remove_column: {"type": "remove_column", "name": str}
                    - For sort: {"type": "sort", "column": str, "ascending": bool}
            
            Returns:
                Dictionary with operation results.
            """
            try:
                file_path = self._dataset_path(user_id, dataset_id)
                
                if not file_path.exists():
                    return {
                        "success": False,
                        "error": f"CSV file not found for dataset '{dataset_id}'"
                    }
                
                # Create backup
                backup_path = self._create_backup(file_path)
                
                # Read current data
                headers, rows = self._read_csv_data(file_path)
                
                # Apply operations
                operations_applied = []
                
                for op in operations:
                    op_type = op.get("type")
                    
                    if op_type == "add_row":
                        new_row = op.get("row", {})
                        rows.append(new_row)
                        operations_applied.append(f"Added row: {new_row}")
                    
                    elif op_type == "update_row":
                        index = op.get("index")
                        if 0 <= index < len(rows):
                            rows[index].update(op.get("row", {}))
                            operations_applied.append(f"Updated row at index {index}")
                        else:
                            operations_applied.append(f"Warning: Invalid index {index}")
                    
                    elif op_type == "delete_row":
                        index = op.get("index")
                        if 0 <= index < len(rows):
                            deleted_row = rows.pop(index)
                            operations_applied.append(f"Deleted row at index {index}")
                        else:
                            operations_applied.append(f"Warning: Invalid index {index}")
                    
                    elif op_type == "rename_column":
                        old_name = op.get("old_name")
                        new_name = op.get("new_name")
                        if old_name in headers:
                            headers[headers.index(old_name)] = new_name
                            for row in rows:
                                if old_name in row:
                                    row[new_name] = row.pop(old_name)
                            operations_applied.append(f"Renamed column '{old_name}' to '{new_name}'")
                    
                    elif op_type == "add_column":
                        col_name = op.get("name")
                        default_value = op.get("default_value", "")
                        if col_name not in headers:
                            headers.append(col_name)
                            for row in rows:
                                row[col_name] = default_value
                            operations_applied.append(f"Added column '{col_name}'")
                    
                    elif op_type == "remove_column":
                        col_name = op.get("name")
                        if col_name in headers:
                            headers.remove(col_name)
                            for row in rows:
                                row.pop(col_name, None)
                            operations_applied.append(f"Removed column '{col_name}'")
                    
                    elif op_type == "sort":
                        col_name = op.get("column")
                        ascending = op.get("ascending", True)
                        if col_name in headers:
                            rows.sort(key=lambda x: x.get(col_name, ""), reverse=not ascending)
                            operations_applied.append(f"Sorted by '{col_name}' ({'ascending' if ascending else 'descending'})")
                
                # Write updated data
                self._write_csv_data(file_path, headers, rows)
                
                # Update metadata
                try:
                    metadata = self._get_metadata(dataset_id, user_id)
                    metadata["columns"] = headers
                    metadata["row_count"] = len(rows)
                    metadata["last_modified"] = datetime.now().isoformat()
                    metadata["backup_file"] = str(backup_path)
                    self._save_metadata(metadata, user_id)
                except FileNotFoundError:
                    # Create metadata if it doesn't exist
                    metadata = {
                        "dataset_id": dataset_id,
                        "original_filename": file_path.name,
                        "stored_filename": file_path.name,
                        "uploaded_at": datetime.now().isoformat(),
                        "size_bytes": file_path.stat().st_size,
                        "columns": headers,
                        "row_count": len(rows),
                        "last_modified": datetime.now().isoformat(),
                        "backup_file": str(backup_path)
                    }
                    self._save_metadata(metadata, user_id)
                
                return {
                    "success": True,
                    "dataset_id": dataset_id,
                    "operations_applied": operations_applied,
                    "new_row_count": len(rows),
                    "backup_file": str(backup_path),
                    "message": f"Applied {len(operations_applied)} operations successfully"
                }
                
            except Exception as e:
                LOGGER.error(f"Error editing CSV file: {e}", exc_info=True)
                return {
                    "success": False,
                    "error": str(e)
                }

        @mcp.tool(tags={self.domain.value})
        def append_to_csv(
            dataset_id: str,
            new_rows: List[Dict[str, Any]],
            user_id: str = "default"
        ) -> Dict[str, Any]:
            """
            Append rows to existing CSV file.
            
            Args:
                dataset_id: The dataset ID to append to
                user_id: User identifier (default: "default")
                new_rows: List of dictionaries representing new rows
            
            Returns:
                Dictionary with append results.
            """
            try:
                file_path = self._dataset_path(user_id, dataset_id)
                
                if not file_path.exists():
                    return {
                        "success": False,
                        "error": f"CSV file not found for dataset '{dataset_id}'"
                    }
                
                # Create backup
                self._create_backup(file_path)
                
                # Read current data
                headers, rows = self._read_csv_data(file_path)
                
                # Append new rows
                rows.extend(new_rows)
                
                # Write updated data
                self._write_csv_data(file_path, headers, rows)
                
                # Update metadata
                try:
                    metadata = self._get_metadata(dataset_id, user_id)
                    metadata["row_count"] = len(rows)
                    metadata["last_modified"] = datetime.now().isoformat()
                    self._save_metadata(metadata, user_id)
                except FileNotFoundError:
                    metadata = {
                        "dataset_id": dataset_id,
                        "original_filename": file_path.name,
                        "stored_filename": file_path.name,
                        "uploaded_at": datetime.now().isoformat(),
                        "size_bytes": file_path.stat().st_size,
                        "columns": headers,
                        "row_count": len(rows),
                        "last_modified": datetime.now().isoformat()
                    }
                    self._save_metadata(metadata, user_id)
                
                return {
                    "success": True,
                    "dataset_id": dataset_id,
                    "rows_added": len(new_rows),
                    "total_rows": len(rows),
                    "message": f"Appended {len(new_rows)} rows successfully"
                }
                
            except Exception as e:
                LOGGER.error(f"Error appending to CSV file: {e}", exc_info=True)
                return {
                    "success": False,
                    "error": str(e)
                }

        @mcp.tool(tags={self.domain.value})
        def merge_csv_files(
            dataset_id1: str,
            dataset_id2: str,
            merge_key: str,
            user_id: str = "default",
            merge_type: str = "inner"
        ) -> Dict[str, Any]:
            """
            Merge two CSV files on a common key column.
            
            Args:
                dataset_id1: First dataset ID
                dataset_id2: Second dataset ID
                merge_key: Column name to merge on
                user_id: User identifier (default: "default")
                merge_type: Type of merge - "inner", "left", "right", or "outer"
            
            Returns:
                Dictionary with merged dataset_id and merge results.
            """
            try:
                file_path1 = self._dataset_path(user_id, dataset_id1)
                file_path2 = self._dataset_path(user_id, dataset_id2)
                
                if not file_path1.exists() or not file_path2.exists():
                    return {
                        "success": False,
                        "error": "One or both CSV files not found"
                    }
                
                # Read both files
                headers1, rows1 = self._read_csv_data(file_path1)
                headers2, rows2 = self._read_csv_data(file_path2)
                
                if merge_key not in headers1 or merge_key not in headers2:
                    return {
                        "success": False,
                        "error": f"Merge key '{merge_key}' not found in one or both files"
                    }
                
                # Create lookup dictionaries
                dict1 = {row[merge_key]: row for row in rows1}
                dict2 = {row[merge_key]: row for row in rows2}
                
                # Perform merge based on type
                merged_rows = []
                if merge_type == "inner":
                    keys = set(dict1.keys()) & set(dict2.keys())
                elif merge_type == "left":
                    keys = set(dict1.keys())
                elif merge_type == "right":
                    keys = set(dict2.keys())
                else:  # outer
                    keys = set(dict1.keys()) | set(dict2.keys())
                
                # Combine headers
                combined_headers = list(headers1)
                for h in headers2:
                    if h != merge_key and h not in combined_headers:
                        combined_headers.append(f"{h}_2")
                
                # Merge rows
                for key in keys:
                    row1 = dict1.get(key, {})
                    row2 = dict2.get(key, {})
                    merged_row = row1.copy()
                    for h in headers2:
                        if h != merge_key:
                            suffix = "_2" if h in headers1 else ""
                            merged_row[f"{h}{suffix}"] = row2.get(h, "")
                    merged_rows.append(merged_row)
                
                # Create new merged dataset
                merged_dataset_id = str(uuid.uuid4())
                merged_filename = f"merged_{dataset_id1[:8]}_{dataset_id2[:8]}.csv"
                merged_path = self._dataset_path(user_id, merged_dataset_id, merged_filename)
                
                self._write_csv_data(merged_path, combined_headers, merged_rows)
                
                # Create metadata
                metadata = {
                    "dataset_id": merged_dataset_id,
                    "original_filename": merged_filename,
                    "stored_filename": merged_filename,
                    "uploaded_at": datetime.now().isoformat(),
                    "size_bytes": merged_path.stat().st_size,
                    "columns": combined_headers,
                    "row_count": len(merged_rows),
                    "created_by": "csv_manipulation_service",
                    "merge_info": {
                        "source_datasets": [dataset_id1, dataset_id2],
                        "merge_key": merge_key,
                        "merge_type": merge_type
                    }
                }
                self._save_metadata(metadata, user_id)
                
                return {
                    "success": True,
                    "merged_dataset_id": merged_dataset_id,
                    "merged_filename": merged_filename,
                    "row_count": len(merged_rows),
                    "columns": combined_headers,
                    "message": f"Merged {len(merged_rows)} rows using {merge_type} join"
                }
                
            except Exception as e:
                LOGGER.error(f"Error merging CSV files: {e}", exc_info=True)
                return {
                    "success": False,
                    "error": str(e)
                }

        @mcp.tool(tags={self.domain.value})
        def filter_csv_rows(
            dataset_id: str,
            filter_conditions: Dict[str, Any],
            user_id: str = "default"
        ) -> Dict[str, Any]:
            """
            Filter rows based on conditions and create new CSV file.
            
            Args:
                dataset_id: Source dataset ID
                filter_conditions: Dictionary with column names as keys and filter values
                    Example: {"status": "active", "age": {"min": 18, "max": 65}}
                    Supports: exact match, min/max for numbers, contains for strings
                user_id: User identifier (default: "default")
            
            Returns:
                Dictionary with filtered dataset_id and filter results.
            """
            try:
                file_path = self._dataset_path(user_id, dataset_id)
                
                if not file_path.exists():
                    return {
                        "success": False,
                        "error": f"CSV file not found for dataset '{dataset_id}'"
                    }
                
                # Read data
                headers, rows = self._read_csv_data(file_path)
                
                # Apply filters
                filtered_rows = []
                for row in rows:
                    matches = True
                    for col, condition in filter_conditions.items():
                        if col not in headers:
                            matches = False
                            break
                        
                        value = row.get(col, "")
                        
                        if isinstance(condition, dict):
                            if "min" in condition:
                                try:
                                    if float(value) < float(condition["min"]):
                                        matches = False
                                        break
                                except (ValueError, TypeError):
                                    matches = False
                                    break
                            if "max" in condition:
                                try:
                                    if float(value) > float(condition["max"]):
                                        matches = False
                                        break
                                except (ValueError, TypeError):
                                    matches = False
                                    break
                            if "contains" in condition:
                                if condition["contains"].lower() not in str(value).lower():
                                    matches = False
                                    break
                        else:
                            if str(value) != str(condition):
                                matches = False
                                break
                    
                    if matches:
                        filtered_rows.append(row)
                
                # Create filtered dataset
                filtered_dataset_id = str(uuid.uuid4())
                filtered_filename = f"filtered_{dataset_id[:8]}.csv"
                filtered_path = self._dataset_path(user_id, filtered_dataset_id, filtered_filename)
                
                self._write_csv_data(filtered_path, headers, filtered_rows)
                
                # Create metadata
                metadata = {
                    "dataset_id": filtered_dataset_id,
                    "original_filename": filtered_filename,
                    "stored_filename": filtered_filename,
                    "uploaded_at": datetime.now().isoformat(),
                    "size_bytes": filtered_path.stat().st_size,
                    "columns": headers,
                    "row_count": len(filtered_rows),
                    "created_by": "csv_manipulation_service",
                    "filter_info": {
                        "source_dataset": dataset_id,
                        "conditions": filter_conditions,
                        "original_row_count": len(rows)
                    }
                }
                self._save_metadata(metadata, user_id)
                
                return {
                    "success": True,
                    "filtered_dataset_id": filtered_dataset_id,
                    "filtered_filename": filtered_filename,
                    "row_count": len(filtered_rows),
                    "original_row_count": len(rows),
                    "filter_conditions": filter_conditions,
                    "message": f"Filtered {len(filtered_rows)} rows from {len(rows)} total rows"
                }
                
            except Exception as e:
                LOGGER.error(f"Error filtering CSV rows: {e}", exc_info=True)
                return {
                    "success": False,
                    "error": str(e)
                }

        @mcp.tool(tags={self.domain.value})
        def transform_csv_columns(
            dataset_id: str,
            transformations: List[Dict[str, Any]],
            user_id: str = "default"
        ) -> Dict[str, Any]:
            """
            Transform CSV columns (add calculated columns, rename, etc.).
            
            Args:
                dataset_id: Source dataset ID
                transformations: List of transformation operations:
                    - {"type": "add_column", "name": str, "formula": str} - Add calculated column
                    - {"type": "rename_column", "old_name": str, "new_name": str}
                    - {"type": "remove_column", "name": str}
                    - {"type": "convert_type", "column": str, "to_type": str} - Convert column type
                user_id: User identifier (default: "default")
            
            Returns:
                Dictionary with transformed dataset_id and transformation results.
            """
            try:
                file_path = self._dataset_path(user_id, dataset_id)
                
                if not file_path.exists():
                    return {
                        "success": False,
                        "error": f"CSV file not found for dataset '{dataset_id}'"
                    }
                
                # Read data
                headers, rows = self._read_csv_data(file_path)
                
                # Apply transformations
                transformations_applied = []
                
                for trans in transformations:
                    trans_type = trans.get("type")
                    
                    if trans_type == "add_column":
                        col_name = trans.get("name")
                        formula = trans.get("formula", "")
                        if col_name not in headers:
                            headers.append(col_name)
                            # Simple formula evaluation (can be enhanced)
                            for row in rows:
                                try:
                                    # Replace column references in formula with actual values
                                    eval_formula = formula
                                    for h in headers:
                                        if h != col_name:
                                            eval_formula = eval_formula.replace(f"{{{h}}}", str(row.get(h, "")))
                                    # Evaluate (basic support - can be enhanced)
                                    row[col_name] = eval(eval_formula.replace("{", "").replace("}", ""))
                                except:
                                    row[col_name] = ""
                            transformations_applied.append(f"Added calculated column '{col_name}'")
                    
                    elif trans_type == "rename_column":
                        old_name = trans.get("old_name")
                        new_name = trans.get("new_name")
                        if old_name in headers:
                            headers[headers.index(old_name)] = new_name
                            for row in rows:
                                if old_name in row:
                                    row[new_name] = row.pop(old_name)
                            transformations_applied.append(f"Renamed '{old_name}' to '{new_name}'")
                    
                    elif trans_type == "remove_column":
                        col_name = trans.get("name")
                        if col_name in headers:
                            headers.remove(col_name)
                            for row in rows:
                                row.pop(col_name, None)
                            transformations_applied.append(f"Removed column '{col_name}'")
                    
                    elif trans_type == "convert_type":
                        col_name = trans.get("column")
                        to_type = trans.get("to_type", "str")
                        if col_name in headers:
                            for row in rows:
                                value = row.get(col_name, "")
                                try:
                                    if to_type == "int":
                                        row[col_name] = int(float(value))
                                    elif to_type == "float":
                                        row[col_name] = float(value)
                                    elif to_type == "str":
                                        row[col_name] = str(value)
                                except (ValueError, TypeError):
                                    pass  # Keep original value if conversion fails
                            transformations_applied.append(f"Converted '{col_name}' to {to_type}")
                
                # Create transformed dataset
                transformed_dataset_id = str(uuid.uuid4())
                transformed_filename = f"transformed_{dataset_id[:8]}.csv"
                transformed_path = self._dataset_path(user_id, transformed_dataset_id, transformed_filename)
                
                self._write_csv_data(transformed_path, headers, rows)
                
                # Create metadata
                metadata = {
                    "dataset_id": transformed_dataset_id,
                    "original_filename": transformed_filename,
                    "stored_filename": transformed_filename,
                    "uploaded_at": datetime.now().isoformat(),
                    "size_bytes": transformed_path.stat().st_size,
                    "columns": headers,
                    "row_count": len(rows),
                    "created_by": "csv_manipulation_service",
                    "transformation_info": {
                        "source_dataset": dataset_id,
                        "transformations": transformations_applied
                    }
                }
                self._save_metadata(metadata, user_id)
                
                return {
                    "success": True,
                    "transformed_dataset_id": transformed_dataset_id,
                    "transformed_filename": transformed_filename,
                    "transformations_applied": transformations_applied,
                    "new_columns": headers,
                    "row_count": len(rows),
                    "message": f"Applied {len(transformations_applied)} transformations"
                }
                
            except Exception as e:
                LOGGER.error(f"Error transforming CSV columns: {e}", exc_info=True)
                return {
                    "success": False,
                    "error": str(e)
                }

        @mcp.tool(tags={self.domain.value})
        def export_csv_as(
            dataset_id: str,
            format: str,
            user_id: str = "default"
        ) -> Dict[str, Any]:
            """
            Export CSV in different formats (JSON, Excel, etc.).
            
            Args:
                dataset_id: Source dataset ID
                format: Export format - "json", "excel", "tsv"
                user_id: User identifier (default: "default")
            
            Returns:
                Dictionary with export file path and format info.
            """
            try:
                file_path = self._dataset_path(user_id, dataset_id)
                
                if not file_path.exists():
                    return {
                        "success": False,
                        "error": f"CSV file not found for dataset '{dataset_id}'"
                    }
                
                # Read data
                headers, rows = self._read_csv_data(file_path)
                
                format_lower = format.lower()
                
                if format_lower == "json":
                    export_path = file_path.with_suffix('.json')
                    with open(export_path, "w", encoding="utf-8") as f:
                        json.dump({"columns": headers, "rows": rows}, f, indent=2)
                
                elif format_lower == "excel":
                    try:
                        import openpyxl
                        export_path = file_path.with_suffix('.xlsx')
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        
                        # Write headers
                        for col_idx, header in enumerate(headers, start=1):
                            ws.cell(row=1, column=col_idx, value=header)
                        
                        # Write rows
                        for row_idx, row_data in enumerate(rows, start=2):
                            for col_idx, header in enumerate(headers, start=1):
                                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                        
                        wb.save(export_path)
                    except ImportError:
                        return {
                            "success": False,
                            "error": "Excel export requires openpyxl library. Install with: pip install openpyxl"
                        }
                
                elif format_lower == "tsv":
                    export_path = file_path.with_suffix('.tsv')
                    with open(export_path, "w", encoding="utf-8", newline="") as f:
                        writer = csv.DictWriter(f, fieldnames=headers, delimiter='\t')
                        writer.writeheader()
                        writer.writerows(rows)
                
                else:
                    return {
                        "success": False,
                        "error": f"Unsupported format: {format}. Supported: json, excel, tsv"
                    }
                
                return {
                    "success": True,
                    "export_path": str(export_path),
                    "format": format,
                    "row_count": len(rows),
                    "message": f"Exported {len(rows)} rows to {format.upper()} format"
                }
                
            except Exception as e:
                LOGGER.error(f"Error exporting CSV: {e}", exc_info=True)
                return {
                    "success": False,
                    "error": str(e)
                }

        @mcp.tool(tags={self.domain.value})
        def validate_csv_structure(
            dataset_id: str,
            expected_schema: Dict[str, str],
            user_id: str = "default"
        ) -> Dict[str, Any]:
            """
            Validate CSV structure against expected schema.
            
            Args:
                dataset_id: Dataset ID to validate
                expected_schema: Dictionary mapping column names to expected types
                    Example: {"name": "str", "age": "int", "score": "float"}
                user_id: User identifier (default: "default")
            
            Returns:
                Dictionary with validation results.
            """
            try:
                file_path = self._dataset_path(user_id, dataset_id)
                
                if not file_path.exists():
                    return {
                        "success": False,
                        "error": f"CSV file not found for dataset '{dataset_id}'"
                    }
                
                # Read data
                headers, rows = self._read_csv_data(file_path)
                
                # Validate schema
                validation_errors = []
                missing_columns = set(expected_schema.keys()) - set(headers)
                extra_columns = set(headers) - set(expected_schema.keys())
                
                if missing_columns:
                    validation_errors.append(f"Missing columns: {list(missing_columns)}")
                
                if extra_columns:
                    validation_errors.append(f"Extra columns: {list(extra_columns)}")
                
                # Validate types for each row
                type_errors = []
                for row_idx, row in enumerate(rows[:10], start=2):  # Check first 10 rows
                    for col_name, expected_type in expected_schema.items():
                        if col_name in headers:
                            value = row.get(col_name, "")
                            try:
                                if expected_type == "int":
                                    int(float(value))
                                elif expected_type == "float":
                                    float(value)
                                elif expected_type == "str":
                                    str(value)
                            except (ValueError, TypeError):
                                type_errors.append(
                                    f"Row {row_idx}, column '{col_name}': expected {expected_type}, got '{value}'"
                                )
                
                is_valid = len(validation_errors) == 0 and len(type_errors) == 0
                
                return {
                    "success": True,
                    "is_valid": is_valid,
                    "validation_errors": validation_errors,
                    "type_errors": type_errors[:10],  # Limit to first 10 errors
                    "columns_found": headers,
                    "columns_expected": list(expected_schema.keys()),
                    "row_count": len(rows),
                    "message": "Validation passed" if is_valid else f"Validation failed: {len(validation_errors)} errors"
                }
                
            except Exception as e:
                LOGGER.error(f"Error validating CSV structure: {e}", exc_info=True)
                return {
                    "success": False,
                    "error": str(e)
                }

        @mcp.tool(tags={self.domain.value})
        def create_csv_from_query(
            dataset_id: str,
            query_logic: Dict[str, Any],
            output_filename: str,
            user_id: str = "default"
        ) -> Dict[str, Any]:
            """
            Create CSV from data query (filter, group, aggregate).
            
            Args:
                dataset_id: Source dataset ID
                query_logic: Query specification:
                    - "select": List of columns to include
                    - "where": Filter conditions (same as filter_csv_rows)
                    - "group_by": Column to group by
                    - "aggregate": Aggregation operations
                        Example: {"sum": "revenue", "count": "orders", "avg": "score"}
                output_filename: Name for output CSV file
                user_id: User identifier (default: "default")
            
            Returns:
                Dictionary with query results and new dataset_id.
            """
            try:
                file_path = self._dataset_path(user_id, dataset_id)
                
                if not file_path.exists():
                    return {
                        "success": False,
                        "error": f"CSV file not found for dataset '{dataset_id}'"
                    }
                
                # Read data
                headers, rows = self._read_csv_data(file_path)
                
                # Apply query logic
                selected_columns = query_logic.get("select", headers)
                where_conditions = query_logic.get("where", {})
                group_by = query_logic.get("group_by")
                aggregate = query_logic.get("aggregate", {})
                
                # Apply filters
                filtered_rows = rows
                if where_conditions:
                    filtered_rows = []
                    for row in rows:
                        matches = True
                        for col, condition in where_conditions.items():
                            if col not in headers:
                                matches = False
                                break
                            value = row.get(col, "")
                            if isinstance(condition, dict):
                                if "min" in condition and float(value) < float(condition["min"]):
                                    matches = False
                                    break
                                if "max" in condition and float(value) > float(condition["max"]):
                                    matches = False
                                    break
                            elif str(value) != str(condition):
                                matches = False
                                break
                        if matches:
                            filtered_rows.append(row)
                
                # Apply grouping and aggregation
                result_rows = []
                if group_by and group_by in headers:
                    # Group by column
                    grouped = {}
                    for row in filtered_rows:
                        key = row.get(group_by, "")
                        if key not in grouped:
                            grouped[key] = []
                        grouped[key].append(row)
                    
                    # Aggregate
                    for key, group_rows in grouped.items():
                        result_row = {group_by: key}
                        for agg_type, col_name in aggregate.items():
                            if col_name in headers:
                                values = [float(r.get(col_name, 0)) for r in group_rows if r.get(col_name, "").replace(".", "").isdigit()]
                                if values:
                                    if agg_type == "sum":
                                        result_row[f"{agg_type}_{col_name}"] = sum(values)
                                    elif agg_type == "avg":
                                        result_row[f"{agg_type}_{col_name}"] = sum(values) / len(values)
                                    elif agg_type == "count":
                                        result_row[f"{agg_type}_{col_name}"] = len(group_rows)
                                    elif agg_type == "min":
                                        result_row[f"{agg_type}_{col_name}"] = min(values)
                                    elif agg_type == "max":
                                        result_row[f"{agg_type}_{col_name}"] = max(values)
                        result_rows.append(result_row)
                    
                    # Update selected columns
                    selected_columns = [group_by] + [f"{agg_type}_{col}" for agg_type, col in aggregate.items()]
                else:
                    # No grouping, just select columns
                    for row in filtered_rows:
                        result_row = {col: row.get(col, "") for col in selected_columns if col in headers}
                        result_rows.append(result_row)
                
                # Create output dataset
                output_dataset_id = str(uuid.uuid4())
                if not output_filename.lower().endswith('.csv'):
                    output_filename = f"{output_filename}.csv"
                
                output_path = self._dataset_path(user_id, output_dataset_id, output_filename)
                self._write_csv_data(output_path, list(selected_columns), result_rows)
                
                # Create metadata
                metadata = {
                    "dataset_id": output_dataset_id,
                    "original_filename": output_filename,
                    "stored_filename": output_filename,
                    "uploaded_at": datetime.now().isoformat(),
                    "size_bytes": output_path.stat().st_size,
                    "columns": list(selected_columns),
                    "row_count": len(result_rows),
                    "created_by": "csv_manipulation_service",
                    "query_info": {
                        "source_dataset": dataset_id,
                        "query_logic": query_logic
                    }
                }
                self._save_metadata(metadata, user_id)
                
                return {
                    "success": True,
                    "output_dataset_id": output_dataset_id,
                    "output_filename": output_filename,
                    "row_count": len(result_rows),
                    "columns": list(selected_columns),
                    "message": f"Created CSV with {len(result_rows)} rows from query"
                }
                
            except Exception as e:
                LOGGER.error(f"Error creating CSV from query: {e}", exc_info=True)
                return {
                    "success": False,
                    "error": str(e)
                }

